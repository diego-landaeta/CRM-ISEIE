"""
Normaliza telefonos de leads ISEIE desde el xlsx oficial.
- Fuente: hoja '2026' del Contactos ISEIE.xlsx
- Toma col B (N de telefono, indice 13) — formato E.164 sin el 1/9 mobile
- Fallback a col A (Telefono, indice 2) si col B esta vacia
- Limpia .0 decimal y espacios
- Match con DB por email O por ultimos 9 digitos del telefono actual
"""
import sys, warnings, json, paramiko
sys.stdout.reconfigure(encoding='utf-8')
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

XLSX = "c:/Users/Diego/Desktop/Proyectos-Carlos/CRM ISEIE/Contactos ISEIE.xlsx"
PROJECT_ID = 10
DRY_RUN = '--dry-run' in sys.argv

def clean_phone(v):
    """Limpia .0 decimal y espacios. Devuelve solo digitos."""
    if v is None: return None
    s = str(v).strip()
    if not s: return None
    if s.endswith('.0'): s = s[:-2]
    digits = ''.join(c for c in s if c.isdigit())
    if len(digits) < 7: return None
    return digits.lstrip('0')

def norm_email(s):
    if not s: return None
    t = str(s).strip().lower()
    if not t or 'no suministrad' in t or '@' not in t: return None
    return t

# 1) Leer xlsx → tabla email/phone → preferred_phone (col B)
print('Leyendo xlsx...')
wb = load_workbook(XLSX, data_only=True)
sh = wb['2026']
records = []
for row in sh.iter_rows(min_row=2, max_row=sh.max_row, values_only=True):
    if not row[0]: continue
    nombre = str(row[0]).strip()
    email = norm_email(row[1])
    phone_a = clean_phone(row[2])  # Telefono
    phone_b = clean_phone(row[13]) if len(row) > 13 else None  # N de telefono
    # Tomar col B si existe, sino col A
    preferred = phone_b or phone_a
    if not email and not preferred: continue
    records.append({'nombre': nombre, 'email': email, 'phone_a': phone_a, 'phone_b': phone_b, 'preferred': preferred})

print(f'Filas xlsx: {len(records)}')
print(f'Con col B preferred: {sum(1 for r in records if r["phone_b"])}')
print(f'Fallback a col A:    {sum(1 for r in records if not r["phone_b"] and r["phone_a"])}')

# 2) SCP el data a server y correr UPDATE
print('\nSubiendo dataset al server...')
data_path = '/tmp/phones_dataset.json'
with open('c:/tmp/phones_dataset.json', 'w', encoding='utf-8') as f:
    json.dump(records, f, ensure_ascii=False)

sys.path.insert(0, 'c:/tmp')
from iseie_ssh import put, run
put(['c:/tmp/phones_dataset.json'], '/tmp')

# 3) Generar script Node que ejecute UPDATE
node_script = '''
import "dotenv/config";
import fs from "fs";
import { query } from "/opt/crm-iseie/src/shared/config/db.js";

const records = JSON.parse(fs.readFileSync("/tmp/phones_dataset.json", "utf-8"));
const DRY_RUN = process.argv.includes("--dry-run");
console.log(`Records a procesar: ${records.length}`);

let updated = 0, sinMatch = 0, sinCambio = 0, samples = [];
for (let i = 0; i < records.length; i++) {
  const r = records[i];
  // Buscar lead por email O por last9 del telefono actual
  let leadRow = null;
  if (r.email) {
    const res = await query(
      `SELECT id, telefono FROM leads WHERE project_id = $1 AND deleted_at IS NULL AND LOWER(email) = $2 LIMIT 1`,
      [10, r.email]
    );
    leadRow = res.rows[0];
  }
  if (!leadRow && r.phone_a) {
    const last9 = r.phone_a.slice(-9);
    const res = await query(
      `SELECT id, telefono FROM leads WHERE project_id = $1 AND deleted_at IS NULL
        AND regexp_replace(telefono, '[^0-9]', '', 'g') LIKE '%' || $2 LIMIT 1`,
      [10, last9]
    );
    leadRow = res.rows[0];
  }
  if (!leadRow && r.phone_b) {
    const last9 = r.phone_b.slice(-9);
    const res = await query(
      `SELECT id, telefono FROM leads WHERE project_id = $1 AND deleted_at IS NULL
        AND regexp_replace(telefono, '[^0-9]', '', 'g') LIKE '%' || $2 LIMIT 1`,
      [10, last9]
    );
    leadRow = res.rows[0];
  }
  if (!leadRow) { sinMatch++; continue; }
  const current = (leadRow.telefono || "").replace(/[^0-9]/g, "");
  if (current === r.preferred) { sinCambio++; continue; }
  if (samples.length < 8) samples.push({ id: leadRow.id, before: current, after: r.preferred });
  if (!DRY_RUN) {
    await query(`UPDATE leads SET telefono = $1, updated_at = NOW() WHERE id = $2`, [r.preferred, leadRow.id]);
  }
  updated++;
  if (i % 1000 === 0 && i > 0) console.log(`  procesados ${i}/${records.length} (${updated} updates)`);
}

console.log(`\\n=== Resumen ===`);
console.log(`Updated:    ${updated}`);
console.log(`Sin cambio: ${sinCambio}`);
console.log(`Sin match:  ${sinMatch}`);
console.log(`\\nSample cambios:`);
for (const s of samples) console.log(`  #${s.id}: ${s.before} -> ${s.after}`);
process.exit(0);
'''
with open('c:/tmp/run_fix_phones.mjs', 'w', encoding='utf-8') as f:
    f.write(node_script)
put(['c:/tmp/run_fix_phones.mjs'], '/opt/crm-iseie')
print('Script subido.')

mode = '--dry-run' if DRY_RUN else ''
print(f'\nEjecutando {"DRY-RUN" if DRY_RUN else "REAL"}...')
rc, out, err = run(f'cd /opt/crm-iseie && node run_fix_phones.mjs {mode} 2>&1 | tail -25', timeout=900)
print(out)
if err: print('STDERR:', err)
